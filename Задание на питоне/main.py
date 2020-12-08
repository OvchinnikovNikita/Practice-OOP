# Овчинников Никита. ПМ-2 (МЕН-293201)

import openpyxl # Для работы с Excel 
from openpyxl.chart import LineChart, Reference, Series # Для построения линейного графика 
import requests # Подключение к сайту
import bs4 # Синтаксический анализ HTML файла

# Ссылка на сайт с информацией о водоёмах Екатеринбурга
link = "https://pogoda1.ru/katalog/sverdlovsk-oblast/temperatura-vody/2-nedeli/#main" 

# Подключение
res = requests.get(link)

# Проверим удалось ли подключение:
try:
	res.raise_for_status()

except Exception as exc: 
	print("Подключение не удалось: %s" % (exc))
 
# Создаём BeautifulSoup объект
bs_object = bs4.BeautifulSoup(res.text, features="html.parser") # В последнем аргументе указываем синтаксический анализатор,
# этого можно не делать, но есть вероятность, что другая система использует иной анализатор, и тогда возможно другое поведение программы
bs_element = bs_object.select("td") # Вытащим все элементы с тегом "td"

# Массив для температур и названий водоёмов
name_and_t_of_the_reservoir = list()

# Пробуем открыть файл 
try: water_file = open("Water.TXT", "w")
except : print("Не удалось открыть файл. Файл занят")

DAY = 7

c = 0
prom = [] # Промежуточный массив

# заполнение массива с именами и температурами
for i in bs_element:
	c += 1
	x = i.getText()
	x = x.replace("()", "") # На момент написания этой части программы на сайта была опечатка, в названии одной реки присутствовало (). Исправил

	if c % 15 != 0:
		water_file.write(x + " ") 
		prom.append(x)
	else:
		water_file.write(x + "\n") # Запишем всё в файл, чтобы потом записать в excel
		c = 0
		prom.append(x)
		name_and_t_of_the_reservoir.append(list(prom))
		prom.clear()

water_file.close()

# Сосчитаем среднюю температуру для каждого дня
average_day_temperature = [0.0 for i in range(DAY * 2)] # Здесь содержатся средние температуры для каждого дня

for i in name_and_t_of_the_reservoir: # Стоит помнить, что мы работаем с матрицей
	counter = -1

	for j in i[1:]:
		counter += 1
		average_day_temperature[counter] += float(j)/17

# Сосчитаем среднюю температуру за две недели 
average_temperature_over_two_weeks = 0

for i in average_day_temperature:
	average_temperature_over_two_weeks += i/(DAY * 2) 

# Начинаем работать с excel
ex_file = openpyxl.load_workbook("Water bodies of Yekaterinburg.XLSX")
sheet = ex_file.active
sheet.title = "Данные о водоёмах" # Меняем название листа

# Считаем ранее записанные данные из файла
with open("Water.TXT") as f:
	read_water = f.readlines()

# Номер строки, массив с обозначениями колонн и счетчик для него
ex_row = 0
ex_column = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R"]
ex_column_count = -1

# Заполняем основную часть таблицы
for i in read_water:
	ex_row += 1

	for j in i.split():
		ex_column_count += 1
		l = ex_column[ex_column_count] + str(ex_row)

		sheet[l] = j

	ex_column_count = -1

# Заполняем средние температуры
ex_column_count = -1
for i in average_day_temperature:
	 ex_column_count += 1
	 l = ex_column[ex_column_count + 2] + "18"

	 sheet[l] = i

# Немного редактирования таблицы для лучшей читаемости
sheet.merge_cells("A18:B18")
sheet.merge_cells("A19:B19")
sheet["A18"] = "Среднее значение для каждого дня:"
sheet["A19"] = "Среднее зачение за две недели:"
sheet["C19"] = average_temperature_over_two_weeks
sheet.column_dimensions["B"].width = 30

# Работа с графиком.
data = Reference(sheet, min_col = 3 , min_row = 18, max_col = 16, max_row = 18) # Определение области считывания данных 
series = Series(data, title = "Среднее значение температуры")         
chart = LineChart() # Линейный график 
chart.title = "График температуры"
chart.append(series)
sheet.add_chart(chart, "C21") 

ex_file.save("Water bodies of Yekaterinburg.XLSX")